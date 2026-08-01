"""Excel (.xlsx) report export.

Shares the ``rows`` / ``(field_key, header_label)`` column-spec shape
used by :mod:`utils.csv_export` and :mod:`utils.pdf` — see that
module's docstring.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGNMENT_RTL = Alignment(horizontal="right", vertical="center")
_BODY_ALIGNMENT_LTR = Alignment(horizontal="left", vertical="center")

_MIN_COLUMN_WIDTH = 10
_MAX_COLUMN_WIDTH = 60


def _autosize_columns(worksheet: Worksheet, column_count: int) -> None:
    """Set each column's width to roughly fit its widest cell's content."""
    for column_index in range(1, column_count + 1):
        letter = get_column_letter(column_index)
        widest = _MIN_COLUMN_WIDTH
        for cell in worksheet[letter]:
            if cell.value is not None:
                widest = max(widest, len(str(cell.value)) + 2)
        worksheet.column_dimensions[letter].width = min(widest, _MAX_COLUMN_WIDTH)


def export_to_excel(
    rows: Sequence[dict[str, Any]],
    columns: Sequence[tuple[str, str]],
    output_path: Path,
    *,
    sheet_title: str = "Report",
    rtl: bool = True,
) -> Path:
    """Write ``rows`` to an ``.xlsx`` file at ``output_path``.

    Args:
        rows: Row data; each dict is keyed by the ``field_key`` from
            ``columns``. A missing key is written as an empty cell.
        columns: Ordered ``(field_key, header_label)`` pairs defining
            which fields to export and what to title each column.
        output_path: Destination file path; parent directories are
            created if needed.
        sheet_title: Worksheet tab name (Excel forbids ``: \\ / ? * [ ]``
            and truncates past 31 characters; both are handled here).
        rtl: Whether the sheet should read right-to-left (column A on
            the right) and body cells right-align — the correct
            orientation for a primarily-Arabic report. Excel's own
            "read Excel content as: [source language]" auto-behavior is
            not something openpyxl can force, so this flag drives it
            explicitly instead.

    Returns:
        ``output_path``, for chaining.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(sheet_title)
    worksheet.sheet_view.rightToLeft = rtl

    field_keys = [key for key, _label in columns]
    headers = [label for _key, label in columns]

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
    worksheet.freeze_panes = "A2"

    body_alignment = _BODY_ALIGNMENT_RTL if rtl else _BODY_ALIGNMENT_LTR
    for row_index, row in enumerate(rows, start=2):
        for column_index, field_key in enumerate(field_keys, start=1):
            cell = worksheet.cell(
                row=row_index, column=column_index, value=row.get(field_key, "")
            )
            cell.alignment = body_alignment

    _autosize_columns(worksheet, len(columns))
    workbook.save(output_path)
    return output_path


def _safe_sheet_title(title: str) -> str:
    """Sanitize a worksheet title to satisfy Excel's naming restrictions."""
    for character in r":\/?*[]":
        title = title.replace(character, "-")
    return title[:31] or "Report"
